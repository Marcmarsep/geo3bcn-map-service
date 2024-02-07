from openpyxl import load_workbook
import os


def parse_xlsx(file_path):
    """
    Parse an XLSX file and extract data into a dictionary.

    Args:
        file_path (str): Path to the XLSX file.

    Returns:
        dict: Data extracted from the XLSX file.
    """
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb['Sheet1']
    data = {}
    for row in ws:
        key = str(row[0].value).strip() if row[0].value else ""
        value = str(row[1].value).strip() if row[1].value else ""
        data[key] = value
    return data


def list_files(directory):
    """
    List all files in a given directory.

    Args:
        directory (str): Directory path to list files from.

    Returns:
        list: A list of file paths.
    """
    return [os.path.join(directory, file) for file in os.listdir(directory)
            if os.path.isfile(os.path.join(directory, file))]


import os


def map_name_list(path,target_map_type):
    """
    List all filenames (without extension) in a given directory, encapsulated in a dictionary with a 'name' key.

    Args:
        directory (str): Directory path to list filenames from.

    Returns:
        list: A list of dictionaries, each containing 'name' as a key and the filename without extension as its value.
    """
    """
    Finds and lists volcanoes that contain a specified file (without extension) 
    in their metadata subdirectories.

    Args:
        volcanoes_dir (str): Path to the 'volcanoes' directory.
        target_file_name (str): Name of the target file without its extension.

    Returns:
        list: A list of names of volcano directories that contain the target file in their metadata subdirectories.
    """
    matching_volcanoes = []

    # Walk through each subdirectory in the volcanoes directory
    for root, dirs, files in os.walk(path):
        # Check if the current directory is a 'metadata' directory
        if os.path.basename(root) == 'metadata':
            # Check each file in the current directory
            for file in files:
                filename_without_extension, _ = os.path.splitext(file)
                # If a file matches the target file name, add the volcano's name to the list
                if filename_without_extension == target_map_type:
                    # Extract the volcano name from the path and add it to the list
                    volcano_name = os.path.basename(os.path.dirname(root))
                    matching_volcanoes.append(volcano_name)
                    break  # Stop checking this metadata directory once a match is found

    return matching_volcanoes


def map_name_list_with_metadata(directory):
    """
    List all filenames (without extension) and create a dict with 'name' key for each in a given directory.

    Args:
        directory (str): Directory path to list filenames from.

    Returns:
        list: A list of dicts with 'name' key for each file.
    """
    files = []
    for file in os.listdir(directory):
        if os.path.isfile(os.path.join(directory, file)):
            filename, _ = os.path.splitext(file)
            files.append({'name': filename})
    return files


def get_map_summary(files):
    """
    Get summary of each map file in the provided list.

    Args:
        files (list): List of file paths.

    Returns:
        list: List of map summaries.
    """
    maps = []
    for file in files:
        map_meta = {}
        wb = load_workbook(filename=file, read_only=True)
        ws = wb['Sheet1']
        data = {}
        filename = os.path.basename(file)

        map_meta["filename"] = filename
        for row in ws:
            key = row[0].value.strip() if row[0].value else ""
            value = row[1].value.strip() if row[1].value else ""
            data[key] = value

            # Extract specific metadata fields
            if key in ["Name", "Volcano Lat", "Volcano Long", "Url"]:
                map_meta[key.lower().replace(" ", "_")] = value

            # Special handling for 'Ll grid point' field
            if key == "Ll grid point (product reference system)":
                lat, _, lng = value.partition(",")
                map_meta["lng"] = lng.strip()
                map_meta["lat"] = lat.strip()

        maps.append(map_meta)
    return maps


def get_types_summary(current_path, volcanoes_path):
    """
    Lists unique filenames found in metadata directories under each volcano directory.

    Args:
        volcanoes_dir (str): Path to the 'volcanoes' directory.

    Returns:
        set: A set of unique filenames found in all metadata directories.
    """
    path = os.path.join(current_path, volcanoes_path)
    unique_filenames = set()

    # Walk through the directory structure starting at volcanoes_dir
    for root, dirs, files in os.walk(path):
        # Check if the current directory is a 'metadata' directory
        if os.path.basename(root) == 'metadata':
            # Extract filename without extension and add to the set
            for file in files:
                filename_without_extension, _ = os.path.splitext(file)
                unique_filenames.add(filename_without_extension)
    return list(unique_filenames)
